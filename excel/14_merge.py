from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합하기
# B2부터 D2까지 합친다
ws.merge_cells("B2:D2") 
ws["B2"].value = "Merged Cell"

# 병합 해제하기.
ws.unmerge_cells("B2:D2")

wb.save("xlsx/sample_merged.xlsx")