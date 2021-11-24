import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 오늘 날짜 정보
ws["A1"] = datetime.datetime.today()
# 실제 엑셀과 같은 수식을 적는다.
ws["A2"] = "=SUM(1,2,3)"
ws["A3"] = "=AVERAGE(1,2,3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("xlsx/sample12.xlsx")