from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook

wb = load_workbook("xlsx/sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# A열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

# 스타일 적용
# 색은 빨갛게 , 이탤릭, 볼드
a1.font = Font(color="FF0000", italic=True, bold=True) 
# 폰트 Arial, 취소선
b1.font = Font(color="CC33FF", name="Arial", strike=True)
# 색은 파랑, 사이즈 20 , 언더라인 
c1.font = Font(color="0000FF", size=20, underline="single")

# 테두리 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), \
                     top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점이 넘는 셀에 대해 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 각 cell에 대해서 가운데 정렬
        cell.alignment = Alignment(horizontal="center",vertical="center")

        # A 번호열은 제외
        if cell.column == 1:
            continue
        # cell.value 가 정수형 데이터 이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            # 배경으로 초록색으로
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid")
            # 폰트의 색상도 변경
            cell.font = Font(color="FF0000")

# 틀고정
# B2기준으로 틀고정
ws.freeze_panes = "B2"

wb.save("xlsx/sample11.xlsx")
