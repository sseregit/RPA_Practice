# 파일 불러오기
from openpyxl import load_workbook
# sample.xlsx 파일에서 wb를 불러온다
wb = load_workbook("sample.xlsx")
# 활성화된 Sheet
ws = wb.active

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column=y).value, end=" ")
    print()    

# cell 갯수를 모를때
for x in range(1, ws.max_row+1): # ws.max_row 최대 row수
    for y in range(1,ws.max_column+1): # ws.max_column 최대 column수
        print(ws.cell(row=x,column=y).value, end=" ")
    print()    
