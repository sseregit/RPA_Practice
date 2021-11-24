from openpyxl import load_workbook

wb = load_workbook("xlsx/sample.xlsx")
ws = wb.active

# A B C => A (B) C D / B C => C D 로 이동
# rows => rows 기준으로 내려감.
# cols => cols 기준으로 옆으로
ws.move_range("B1:C11", rows=0, cols=1)
# 비어있는 B에 입력
ws["B1"].value = "국어"

# A B C => A (B) C D / B C => C D 로 이동
# rows => rows 기준으로 내려가면서 값이 있을경우 덮어씌어버림
# cols => cols 기준으로 옆으로
ws.move_range("C1:C11", rows=5, cols=-1)
# 비어있는 B에 입력
ws["B1"].value = "국어"

wb.save("xlsx/sample9.xlsx")
