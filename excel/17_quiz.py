from openpyxl import Workbook
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))
ws.append(("1",10,8,5,14,26,12))
ws.append(("2",7,3,7,15,24,18))
ws.append(("3",9,5,8,8,12,4))
ws.append(("4",7,8,7,17,21,18))
ws.append(("5",7,8,7,16,25,15))
ws.append(("6",3,5,8,8,17,0))
ws.append(("7",4,9,10,16,27,18))
ws.append(("8",6,6,6,15,19,17))
ws.append(("9",10,10,9,19,30,19))
ws.append(("10",9,8,8,20,25,20))

# 총점과 성적정보
ws["H1"] = "총점"
ws["I1"] = "성적정보"

# 퀴즈2 점수 10으로 수정
quizs2 = ws["D"]
for quiz2 in quizs2[1:]:
    quiz2.value = 10

# H열에 총점(SUM) / I열에 등급
for row in tuple(ws.rows)[1:]:
    idx = row[2].coordinate[1:]
    ws["H"+idx] = f"=SUM(B{idx}:G{idx})"
    result = 0
    for cell in row[1:-2]:
        result += cell.value

    if result >= 90:
        ws["I"+idx] = "A"
    elif result >= 80:
        ws["I"+idx] = "B"
    elif result >= 70:
        ws["I"+idx] = "C"
    else:
        ws["I"+idx] = "D"

    if 5 > row[1].value:
        ws["I"+idx] = "F"
        
wb.save("xlsx/scores.xlsx")
