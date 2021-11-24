from openpyxl import load_workbook

wb = load_workbook("xlsx/sample12.xlsx")
ws = wb.active

# values value정보를 바로가져온다.
# 수식은 수식 그대로 가져옴.
for row in ws.values:
    for cell in row:
        print(cell)

# data_only=True / 셀의 수식이 아닌 데이터를 가져온다
wb = load_workbook("xlsx/sample12.xlsx", data_only=True)
ws = wb.active

# values value정보를 바로가져온다.
# evaluate 되지 않은 상태의 데이터는 None
# openpyxl로 수식을 만들었을시 수식의 값은 저장되지않음. 따로 저장해야함.
# None이 나오는게 맞음. 다시 저장해서 엑셀안에 수식의 값을 저장해야함.
for row in ws.values:
    for cell in row:
        print(cell)

wb.save("xlsx/sample13.xlsx")