from openpyxl import Workbook

wb = Workbook()
# 새로운 Sheet 생성
ws = wb.create_sheet()
# Sheet 이름변경
ws.title = "MySheet"
# RGB 형식으로 색상을 지정한다.
ws.sheet_properties.tabColor = "ff99cc"

# sheet를 생성하면서 바로 title을 줄수있다.
ws1 = wb.create_sheet("YourSheet")
# 2번째 index에 Sheet 생성
ws2 = wb.create_sheet("NewSheet",2)

# 위처럼 변수가아니라 dict 형태로 Sheet에 접근이 가능하다.
new_ws = wb["NewSheet"]
# 현재 있는 Sheet 모든 이름을 가져온다.
print(wb.sheetnames)

# Sheet 복사 
new_ws["A1"] = "Test"
# Workbook 내에 있는 new_ws Sheet를 복사해서 target에 넣어줌
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheep"

wb.save("sample.xlsx")

wb.close()