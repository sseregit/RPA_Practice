from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Excel"
# [index] 위치 값을 넣는다.
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# A1 셀의 정보
print(ws["A1"])
# A1 셀의 값 / 값이 없다면 None return
print(ws["A1"].value)

# A1 cell과 같다. index가 1부터 시작된다.
ws.cell(row=1,column=1).value
ws.cell(row=1,column=2).value
# ws["C1"] = 10 과 같다 값을 바로 넣음 value
c1 = ws.cell(row=1,column=3,value="10")
print(c1.value)

from random import *

# 반복문을 이용한 랜덤 숫자 채우기
for x in range(1,11): # row
    for y in range(1,11): # column
        ws.cell(row=x, column=y, value=randint(0,100))

wb.save("sample.xlsx")