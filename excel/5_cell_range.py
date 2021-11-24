from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# append 데이터 한줄씩 넣는다 
ws.append(("번호","영어","수학"))

for i in range(1,11):
    ws.append((i,randint(0,100),randint(0,100)))

# B column을 가져온다.
col_B = ws["B"]
for cell in col_B:
    print(cell.value)

# 다중column 가져오기 B 와 C column 을 가져오기.
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value)

# 첫번째 row만 가져오기
row_title = ws[1]
for cell in row_title:
    print(cell.value)

# 다중 row 가져오기 2 ~ 6째 까지 가져오기.
row_range = ws[2:6]
for rows in row_range:
    for row in rows:
        print(row.value,end=" ")
    print()

# 2번째 줄부터 마지막까지.
row_range = ws[2:ws.max_row]
for rows in row_range:
    for row in rows:
        print(row.value,end=" ")
    print()

# 특정열마다 조건이 다를경우
from openpyxl.utils.cell import coordinate_from_string
# 2번째 줄부터 마지막까지.
row_range = ws[2:ws.max_row]
for rows in row_range:
    for row in rows:
        # cell의 좌표정보를 가져온다.
        print(row.coordinate,end=" ")
        # 좌표 (column, index) 를 return 해준다.
        xy = coordinate_from_string(row.coordinate)
        print(xy, end= " ")
    print()

# 전체 rows를 가져온다. 한 row씩 모든 row를 가져옴
rows = tuple(ws.rows)
for row in rows:
    print(row)

# 전체 columns를 가져온다. 한 column씩 모든 column을 가져온다.
columns = tuple(ws.columns)
for column in columns:
    print(column[0].value)

# 전체 row를 가져온다.
for row in ws.iter_rows():
    print(row)

# 전체 row중 1 ~ 5째 row까지 가져온다. 2열 ~ 3열까지
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row)

# 전체 column를 가져온다.
for column in ws.iter_cols():
    print(column)

# 전체 column를 가져온다.
for column in ws.iter_cols(min_row=1, max_row=5, min_col=2, max_col=3):
    print(column)

wb.save("sample.xlsx")