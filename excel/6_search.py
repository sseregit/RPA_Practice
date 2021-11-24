from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# 값을 찾아서 조건걸기.
for row in ws.iter_rows(min_row=2):
    if row[1].value > 80:
        print(row[0].value,"번 학생은 영어 천재")

# 값을 찾아서 변경하기.
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample.xlsx")